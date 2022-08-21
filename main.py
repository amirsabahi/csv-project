import os
import csv
import pathlib
import time
import argparse
from openpyxl import load_workbook, Workbook

'''
Read CSV
'''


def open_csv(csv_path):
    rows = []
    with open(csv_path, 'r') as csv_file:
        csv_content = csv.reader(csv_file)
        try:
            for row in csv_content:
                rows.append(row)
        except Exception as e:
            log('error', e)

    return rows


''' Open Exccel '''


def open_excel(xlsx_path):
    try:
        data = load_workbook(xlsx_path, read_only=True)
    except Exception as e:
        log('error', e)
        return False
    return data


# customized for our project we get only columns with date
def process_xlsx(data, sheet_number=0):
    dates = []
    for cell in data[data.sheetnames[sheet_number]]:
        datetime = cell[1:3]
        if datetime[0].value is None or datetime[1].value is None:
            continue
        try:
            time_cell = datetime[1].value
            date_cell = time.mktime(
                time.strptime(f'{datetime[0].value.year}-{datetime[0].value.month}-{datetime[0].value.day} {time_cell}',
                              '%Y-%m-%d %H:%M:%S'))
            dates.append(date_cell)
        except Exception as e:
            # log('error', e)
            continue

    return dates


''' 
Get start and end time from the given rows
Return a list of Unix timestamp. First is start timme and the second list item is the end time
'''


def get_times(rows):
    # Convert the date time to timestamp
    start_time = time.mktime(time.strptime(f'{rows[1][1:2][0]} {rows[1][2:3][0]}', '%d.%m.%Y %H:%M:%S'))
    end_time = time.mktime(time.strptime(f'{rows[-1][1:2][0]} {rows[-1][2:3][0]}', '%d.%m.%Y %H:%M:%S'))
    return [start_time, end_time]


def search_cache(search_timestamp, items):
    # not an optimal search
    index = 0

    for item in items:

        if item[0] <= search_timestamp <= item[1]:
            return index
        index += 1
    return None


def log(log_type, message):
    if log_type == 'info':
        print(f"\033[92m {message}\033[00m")
        return None
    if log_type == 'error':
        print(f"\033[91m {message}\033[00m")
        return None
    if log_type == 'warning':
        print(f"\033[93m {message}\033[00m")
        return None
    print(f"\n {message} \n")
    return None


''' 
Main command
'''


def command():
    log('info', 'Started processing the documents.')
    parser = argparse.ArgumentParser(description='A simple command to detect video locations based on the'
                                                 ' given date and time. For current directory use . only.')
    parser.add_argument('-d', required=True, help="Provide the folder url.", default='.', metavar='path')
    args = parser.parse_args()
    try:
        dirs = os.scandir(path=args.d)
    except FileNotFoundError as e:
        log('error', e)
        log('warning', 'Exit!')
        exit(-1)
    main_xlsx_file = None
    data_file_dir = None
    for dir in dirs:
        # Get main excel file
        if pathlib.Path(dir).suffix == '.xlsx':
            main_xlsx_file = pathlib.Path(dir)
        if dir.is_dir():
            data_file_dir = pathlib.Path(dir)
    if main_xlsx_file is None or data_file_dir is None:
        log('error', 'Error: No Excel file found!')
        log('warning', 'Exit!\n')
        exit(-1)

    cache = {}

    data_dirs = os.scandir(data_file_dir)
    data_dirs_uri = []
    for dir_name in data_dirs:
        name = dir_name.name
        if name[:4] == "DATA":
            data_dirs_uri.append(name)
            # Get csv files only
            files = os.listdir(dir_name)

            files.sort()
            for file in files:
                if file[-3:] == "csv":
                    data_csv_dir = f'{data_file_dir}/{dir_name.name}'
                    relative_path = f'{data_file_dir}/{dir_name.name}/{file}'
                    # open CSV
                    rows = open_csv(relative_path)

                    # Get first and last date time and store in cache
                    if rows:
                        times = get_times(rows)
                        cache[data_csv_dir] = times

    # open the main Excel and read

    main_excel = open_excel(main_xlsx_file)
    main_dates = process_xlsx(main_excel)

    cache_values = cache.values()
    cache_keys = list(cache.keys())

    wbook = Workbook()
    wsheet = wbook.create_sheet("Links", 0)
    i = 1
    for timestamp in main_dates:
        wsheet[f'A{i}'].value = os.path.abspath(cache_keys[search_cache(timestamp, cache_values)])
        i += 1

    wbook.save('path.xlsx')
    log('info', '\nSuccess: File path.xlsx Saved.\n')
    log(None, '\nGoodbye!\n')
    exit(1)


if __name__ == '__main__':
    command()
